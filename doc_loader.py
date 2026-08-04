"""Robust document loading with local-only PDF OCR.

SimpleDirectoryReader and pypdf only extract embedded PDF text. Scanned or
image-only PDFs have pixels but no text layer, so those readers correctly
return empty pages. This loader centralizes the fallback chain used by every
engine:

SimpleDirectoryReader -> PyMuPDF text -> pypdf text -> PyMuPDF OCR with Tesseract
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Iterable


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _data_like(cell: str) -> bool:
    """True for cell values that belong in records, not header labels.

    Pure numbers, ISO dates/datetimes, and URLs are record values.  A candidate
    header row full of these is really the first data row.
    """
    return bool(
        re.match(r"^https?://", cell, flags=re.I)
        or re.match(r"^\d{4}-\d{2}-\d{2}", cell)
        or re.fullmatch(r"\(?-?[\d,]*\.?\d+\)?%?", cell)
    )


def _header_row_index(table: list[tuple]) -> int | None:
    """Find the real header in a report-style sheet rather than assuming row one.

    XLSX exports commonly start with titles, notes, or blank rows.  A usable
    header is mostly text labels and is followed by a dense block of records.
    """
    best: tuple[float, int] | None = None
    for index, row in enumerate(table[:10]):
        cells = [_cell_text(value) for value in row]
        labels = [cell for cell in cells if cell]
        text_labels = [cell for cell in labels if not cell.replace(".", "", 1).replace("-", "", 1).isdigit()]
        # Repeated labels (for example, several "Total" columns) are normal
        # in report workbooks with stacked headers.  They are disambiguated by
        # _clean_headers using the preceding header rows.
        if len(labels) < 2 or len(text_labels) < 2:
            continue
        width = len(labels)
        following = table[index + 1:index + 13]
        densities = [sum(bool(_cell_text(value)) for value in candidate) for candidate in following]
        dense_rows = sum(density >= max(2, width * 0.45) for density in densities)
        if dense_rows < 2:
            continue
        score = len(text_labels) * 4 + dense_rows * 3 + min(width, 12)
        if best is None or score > best[0]:
            best = (score, index)
    if best is None:
        return None
    # Sparse trackers can score a record row above the real header because the
    # density threshold scales with the candidate's own width.  A row holding
    # dates/URLs/numbers under a fully-labelled text row is data, not a header.
    index = best[1]
    while index > 0:
        current = [_cell_text(value) for value in table[index] if _cell_text(value)]
        above = [_cell_text(value) for value in table[index - 1] if _cell_text(value)]
        if (
            sum(_data_like(cell) for cell in current) >= 2
            and above
            and not any(_data_like(cell) for cell in above)
            and len(above) >= max(2, len(current) - 2)
        ):
            index -= 1
            continue
        break
    return index


def _clean_headers(header_rows: list[tuple]) -> list[str]:
    seen: dict[str, int] = {}
    headers = []
    width = max(len(row) for row in header_rows)
    # Excel represents merged cells only in their left-most position. Forward
    # fill each header level so a group label applies to its whole span.
    expanded_rows = []
    for row in header_rows:
        expanded = []
        last_value = ""
        for index in range(width):
            value = _cell_text(row[index] if index < len(row) else None)
            if value:
                last_value = value
            expanded.append(last_value)
        expanded_rows.append(expanded)
    for index in range(width):
        parts = []
        for row in expanded_rows:
            value = row[index]
            if value and value not in parts:
                parts.append(value)
        base = " | ".join(parts) or f"Column {index + 1}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def _key_value_layout(table: list[tuple]) -> tuple[list[str] | None, int]:
    """Recognize headerless summary sheets shaped as label/value pairs.

    Report cards such as "Total people contacted | 40" have no header row at
    all, so they were previously skipped and invisible to every query engine.
    """
    populated = []
    for row in table:
        filled = [(column, _cell_text(value)) for column, value in enumerate(row) if _cell_text(value)]
        if filled:
            populated.append(filled)
    if len(populated) < 3:
        return None, 0
    pairs = 0
    for filled in populated:
        if not (1 <= len(filled) <= 2 and filled[0][0] == 0 and filled[-1][0] <= 1):
            return None, 0
        if _data_like(filled[0][1]):
            return None, 0
        if len(filled) == 2:
            pairs += 1
    if pairs < 3:
        return None, 0
    return ["Item", "Value"], 0


def _merge_subheader_row(
    table: list[tuple], header_start: int, header_index: int, headers: list[str]
) -> tuple[list[str], int] | None:
    """Fold a cross-tab label row below the header into the header itself.

    A matrix like "Semester" spanning three columns with a following row of
    "Summer C 2025 / Fall 2025 / Spring 2026" and X marks underneath needs the
    labels merged into the headers; otherwise the labels are stored as a data
    row and the X-mark columns are indistinguishable.
    """
    bases: dict[str, int] = {}
    for header in headers:
        base = re.sub(r"_\d+$", "", header)
        bases[base] = bases.get(base, 0) + 1
    ambiguous = {
        column for column, header in enumerate(headers)
        if bases[re.sub(r"_\d+$", "", header)] > 1 or header.startswith("Column ")
    }
    if not ambiguous:
        return None
    for index in range(header_index + 1, min(header_index + 4, len(table))):
        cells = [_cell_text(value) for value in table[index]]
        filled = [column for column, cell in enumerate(cells) if cell]
        if not filled:
            continue
        if not set(filled) <= ambiguous or any(_data_like(cells[column]) for column in filled):
            return None
        following = table[index + 1:index + 7]
        dense = sum(
            sum(bool(_cell_text(value)) for value in row) >= 2 for row in following
        )
        if dense < 2:
            return None
        merged = _clean_headers([*table[header_start:header_index + 1], table[index]])
        return merged, index + 1
    return None


def _total_chars(docs: Iterable[Document]) -> int:
    return sum(len(doc.text or "") for doc in docs)


def _sheet_merged_ranges(file_path: Path) -> dict[str, list[tuple[int, int, int, int]]]:
    """Merged-cell ranges per worksheet, read straight from the workbook XML.

    openpyxl's read-only mode does not expose merged cells, and a full load of
    a large workbook is slow, so the <mergeCells> entries are read directly.
    Returns {sheet name: [(min_col, min_row, max_col, max_row), ...]} 1-indexed.
    """
    import zipfile
    from xml.etree import ElementTree

    from openpyxl.utils.cell import range_boundaries

    main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    result: dict[str, list[tuple[int, int, int, int]]] = {}
    try:
        with zipfile.ZipFile(file_path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            targets = {item.get("Id"): item.get("Target", "") for item in rels}
            for sheet in workbook.iter(f"{main}sheet"):
                target = targets.get(sheet.get(f"{rel}id"), "")
                member = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
                sheet_xml = ElementTree.fromstring(archive.read(member))
                result[sheet.get("name")] = [
                    range_boundaries(merge.get("ref"))
                    for merge in sheet_xml.iter(f"{main}mergeCell")
                    if merge.get("ref")
                ]
    except Exception:
        return {}
    return result


def _get_config_value(name: str, default=None):
    try:
        from config import Config

        return getattr(Config, name, default)
    except Exception:
        return default


def _ocr_dpi() -> int:
    configured = os.getenv("PDF_OCR_DPI") or str(_get_config_value("PDF_OCR_DPI", 200))
    try:
        return max(120, min(int(configured), 300))
    except ValueError:
        return 200


def _tessdata_path() -> Path | None:
    configured = os.getenv("TESSDATA_PREFIX")
    candidates = [
        Path(configured) if configured else None,
        Path(r"C:\Program Files\Tesseract-OCR\tessdata"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tessdata"),
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def _page_documents(
    page_text: list[tuple[int, str]],
    file_path: Path,
    extraction_method: str,
) -> list[Document]:
    from llama_index.core.schema import Document

    docs = []
    for page_number, text in page_text:
        cleaned = (text or "").strip()
        if not cleaned:
            continue
        docs.append(
            Document(
                text=cleaned,
                metadata={
                    "file_name": file_path.name,
                    "file_path": str(file_path),
                    "page": page_number,
                    "extraction_method": extraction_method,
                },
            )
        )
    return docs


def _cache_table(
    rows: Iterable[tuple],
    file_path: Path,
    *,
    sheet_name: str,
    formula_rows: Iterable[tuple] | None = None,
    merged_ranges: list[tuple[int, int, int, int]] | None = None,
) -> None:
    """Store the primary tabular region exactly in SQLite."""
    table = list(rows)
    header_index = _header_row_index(table)
    if header_index is None:
        headers, data_start = _key_value_layout(table)
        if headers is None:
            return
    else:
        # Keep up to three contiguous header rows so a report matrix such as
        # "Photovoltaic → Capacity MW → Residential" becomes one usable field.
        header_start = header_index
        while header_start > 0 and header_index - header_start < 2:
            if not any(_cell_text(value) for value in table[header_start - 1]):
                break
            header_start -= 1
        headers = _clean_headers(table[header_start:header_index + 1])
        data_start = header_index + 1
        merged = _merge_subheader_row(table, header_start, header_index, headers)
        if merged is not None:
            headers, data_start = merged
    data_indices = []
    blank_run = 0
    for index, row in enumerate(table[data_start:], start=data_start):
        populated = [_cell_text(value) for value in row if _cell_text(value)]
        if not populated:
            blank_run += 1
            if blank_run >= 3:
                break
            continue
        blank_run = 0
        # Report footnotes are often single-cell rows around the data block —
        # a long note sentence or a bare link line. They are metadata, not records.
        if len(populated) == 1 and populated[0].lower().startswith(("http://", "https://", "www.")):
            continue
        if data_indices and len(populated) == 1 and len(populated[0]) >= 80:
            break
        data_indices.append(index)
    # Excel stores a merged cell's value only in its top-left anchor, so group
    # labels spanning several records (a merged "Specialization" column) read
    # as blanks. Propagate anchors into the selected data rows only — header
    # rows are handled by _clean_headers, and footer merges stay untouched.
    overrides: dict[int, dict[int, object]] = {}
    data_set = set(data_indices)
    for min_col, min_row, max_col, max_row in merged_ranges or []:
        anchor_row = table[min_row - 1] if min_row - 1 < len(table) else ()
        anchor = anchor_row[min_col - 1] if min_col - 1 < len(anchor_row) else None
        if not _cell_text(anchor):
            continue
        for row_index in range(min_row - 1, max_row):
            if row_index not in data_set:
                continue
            for column_index in range(min_col - 1, max_col):
                if column_index < len(headers):
                    overrides.setdefault(row_index, {})[column_index] = anchor

    def _filled(index: int) -> tuple:
        row = table[index]
        cells = overrides.get(index)
        if not cells:
            return row
        widened = list(row) + [None] * (max(cells) + 1 - len(row))
        for column_index, value in cells.items():
            if not _cell_text(widened[column_index]):
                widened[column_index] = value
        return tuple(widened)

    materialized_rows = [_filled(index) for index in data_indices]
    formula_table = list(formula_rows) if formula_rows is not None else None
    materialized_formulas = (
        [formula_table[index] for index in data_indices if index < len(formula_table)]
        if formula_table is not None else None
    )
    from spreadsheet_store import index_sheet

    index_sheet(file_path, sheet_name, headers, materialized_rows, materialized_formulas)


def _ingest_csv(file_path: Path) -> None:
    import csv

    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        _cache_table(csv.reader(handle), file_path, sheet_name="CSV")


def _ingest_xlsx(file_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires openpyxl. Install requirements.txt and restart the app.") from exc

    # Values answer financial questions; formulas expose model logic when requested.
    values_book = load_workbook(file_path, read_only=True, data_only=True)
    formulas_book = load_workbook(file_path, read_only=True, data_only=False)
    merged_by_sheet = _sheet_merged_ranges(file_path)
    try:
        for values_sheet, formulas_sheet in zip(values_book.worksheets, formulas_book.worksheets):
            _cache_table(
                values_sheet.iter_rows(values_only=True),
                file_path,
                sheet_name=values_sheet.title,
                formula_rows=formulas_sheet.iter_rows(values_only=True),
                merged_ranges=merged_by_sheet.get(values_sheet.title),
            )
    finally:
        values_book.close()
        formulas_book.close()


def ingest_spreadsheet(file_path: str | Path, force: bool = False) -> int:
    """Parse a workbook once and make its exact rows immediately queryable."""
    file_path = Path(file_path)
    from spreadsheet_store import (
        begin_ingestion,
        cached_row_count,
        fail_ingestion,
        finish_ingestion,
        is_ready,
    )

    if not force and is_ready(file_path):
        return cached_row_count(file_path)
    begin_ingestion(file_path)
    try:
        if file_path.suffix.lower() == ".csv":
            _ingest_csv(file_path)
        elif file_path.suffix.lower() == ".xlsx":
            _ingest_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported spreadsheet type: {file_path.suffix}")
        count = finish_ingestion(file_path)
        print(f"[spreadsheet_store] Ready: {file_path.name} ({count} rows)")
        return count
    except Exception as exc:
        fail_ingestion(file_path, exc)
        raise


def _load_spreadsheet_documents(file_path: Path) -> list[Document]:
    from llama_index.core.schema import Document
    from spreadsheet_store import semantic_records

    ingest_spreadsheet(file_path)
    return [Document(text=item["text"], metadata=item["metadata"]) for item in semantic_records(file_path)]


def _load_with_pymupdf_text(file_path: Path) -> list[Document]:
    import fitz

    pdf = fitz.open(str(file_path))
    try:
        pages = [(i + 1, page.get_text("text")) for i, page in enumerate(pdf)]
    finally:
        pdf.close()
    return _page_documents(pages, file_path, "pymupdf")


def _load_with_pypdf_text(file_path: Path) -> list[Document]:
    import pypdf

    reader = pypdf.PdfReader(str(file_path))
    pages = [(i + 1, page.extract_text() or "") for i, page in enumerate(reader.pages)]
    return _page_documents(pages, file_path, "pypdf")


def _load_with_pymupdf_ocr(file_path: Path) -> list[Document]:
    """Use PyMuPDF's Tesseract bridge for local OCR."""
    import fitz

    tessdata = _tessdata_path()
    if not tessdata:
        raise RuntimeError(
            "Tesseract tessdata was not found. Install Tesseract OCR or set "
            "TESSDATA_PREFIX to the tessdata directory."
        )

    pdf = fitz.open(str(file_path))
    pages: list[tuple[int, str]] = []
    try:
        for i, page in enumerate(pdf):
            textpage = page.get_textpage_ocr(
                flags=0,
                language="eng",
                dpi=_ocr_dpi(),
                full=True,
                tessdata=str(tessdata),
            )
            pages.append((i + 1, page.get_text("text", textpage=textpage)))
    finally:
        pdf.close()
    return _page_documents(pages, file_path, "pymupdf_ocr")


def load_documents(file_path: str | Path) -> list[Document]:
    """Load a file into LlamaIndex Documents, OCRing scanned PDFs locally."""
    file_path = Path(file_path)
    fname = file_path.name
    is_pdf = fname.lower().endswith(".pdf")

    if file_path.suffix.lower() == ".csv":
        docs = _load_spreadsheet_documents(file_path)
        if not docs:
            raise ValueError(f"No rows were found in '{fname}'.")
        return docs
    if file_path.suffix.lower() == ".xlsx":
        docs = _load_spreadsheet_documents(file_path)
        if not docs:
            raise ValueError(f"No readable rows were found in '{fname}'.")
        return docs

    from llama_index.core import SimpleDirectoryReader

    docs = SimpleDirectoryReader(input_files=[str(file_path)]).load_data()
    total_chars = _total_chars(docs)
    print(f"[loader] Loaded '{fname}': {len(docs)} docs, {total_chars} chars")

    if total_chars == 0 and is_pdf:
        print(f"[loader] Empty PDF text for '{fname}', trying PyMuPDF text...")
        try:
            docs = _load_with_pymupdf_text(file_path)
            total_chars = _total_chars(docs)
            if total_chars:
                print(
                    f"[loader] PyMuPDF extracted '{fname}': "
                    f"{len(docs)} pages, {total_chars} chars"
                )
        except Exception as exc:
            print(f"[loader] PyMuPDF text failed for '{fname}': {exc}")

    if total_chars == 0 and is_pdf:
        print(f"[loader] Trying pypdf text extraction for '{fname}'...")
        try:
            docs = _load_with_pypdf_text(file_path)
            total_chars = _total_chars(docs)
            if total_chars:
                print(
                    f"[loader] pypdf extracted '{fname}': "
                    f"{len(docs)} pages, {total_chars} chars"
                )
        except Exception as exc:
            print(f"[loader] pypdf text failed for '{fname}': {exc}")

    if total_chars == 0 and is_pdf:
        print(f"[loader] PDF '{fname}' appears image-only, trying local OCR...")
        try:
            docs = _load_with_pymupdf_ocr(file_path)
            total_chars = _total_chars(docs)
            if total_chars:
                print(
                    f"[loader] Local OCR extracted '{fname}': "
                    f"{len(docs)} pages, {total_chars} chars"
                )
        except Exception as exc:
            raise RuntimeError(
                f"Could not extract text from '{fname}'. It looks like an image-only "
                "or scanned PDF, and local OCR failed. Install Tesseract OCR and "
                "make sure both tesseract.exe and tessdata are available. On "
                "Windows: winget install UB-Mannheim.TesseractOCR, then set "
                "TESSDATA_PREFIX to C:\\Program Files\\Tesseract-OCR\\tessdata. "
                f"Original OCR error: {exc}"
            ) from exc

    if is_pdf and total_chars == 0:
        raise ValueError(
            f"No readable text was found in '{fname}'. The PDF may be blank, "
            "encrypted, or a scan with unreadable images."
        )

    if docs and docs[0].text.strip():
        print(f"[loader] Preview for '{fname}': {repr(docs[0].text[:200])}")

    return docs
